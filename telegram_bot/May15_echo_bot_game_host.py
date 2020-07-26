from openpyxl import Workbook
import openpyxl
import os.path

def compute(parts, action):
    if os.path.exists("gamers_list.xlsx"):
        print("File already exists. Modifying...")
        # wb = Workbook()
        # book = openpyxl.load_workbook("gamers_list.xlsx")
        wb = openpyxl.load_workbook("gamers_list.xlsx")
        ws = wb.active
        # sheet = book.active

    else:
        print("File doesnt exist yet. Creating")
        wb = Workbook()
        #grab the active worksheet
        ws = wb.active
        #Data can be assigned directly to cells
        # ws['A1'] = 42
        ws.append(["Name", "Game", "Venue", "Date"])

    if action == "add":
        print("Adding name")
        data = [parts[1], parts[2], parts[3], parts[4]]
        ws.append(data)
    else:
        print("Deleting Name")
        count = 0
        while True:
            name = ws[f'A{count+2}']
            if name.value == None:
                print("Not in the list")
                return -1
                break
            if name.value == parts[1]:
                print(f"Found name in row {count+2}")
                ws.delete_rows(count+2, 1)
                break
            else:
                count += 1
                continue
    wb.save("gamers_list.xlsx")


def identify_msg(msg):
    event_details = "Frisbee - 8pm Friday 17th May @ BFS"

    help_msg = "For details on upcoming games..."

    question_phrases = ["game on",
                        "match happening",
                        "plan",
                        "event"]

    for phrase in question_phrases:
        if phrase in msg:
            return event_details

    parts = msg.split(',')
    if parts[0] == "In":
        action = "add"
        compute(parts, action)
        output = f"{parts[1]}, You are added to the list!"
    elif parts[0] == "Out":
        action = "remove"
        compute(parts, action)
        if compute(parts, action) == -1:
            output = f"{parts[1]}, NOT ON THE LIST!"
        else:
            output = f"{parts[1]}, You are removed from the list."
    else:
        output = help_msg

    return output
    

if __name__ == "__main__":


    msg = "In, Karan, Frisbee, BFS, 17th, "
    result = identify_msg(msg)
    print(result)